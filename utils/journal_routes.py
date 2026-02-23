import atexit
import base64
import io
import json
import re
import secrets
from datetime import date, datetime, timezone
from urllib.parse import urlparse

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

from flask import Response, flash, jsonify, redirect, render_template, request, send_file, stream_with_context, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func, or_

from utils.journal_models import (
    ATTENDANCE_STATUS_ABSENT,
    ATTENDANCE_STATUS_EXCUSED,
    ATTENDANCE_STATUS_PRESENT,
    ATTENDANCE_STATUSES,
)
from utils.journal_realtime import RealtimeEventBus
from utils.journal_tunnel import JournalTunnelManager, is_local_request

try:
    import segno
except Exception:
    segno = None

WEEK_PARITY_OPTIONS = ("I", "II")
DAY_OPTIONS = (
    {"id": 1, "name": "Понедельник"},
    {"id": 2, "name": "Вторник"},
    {"id": 3, "name": "Среда"},
    {"id": 4, "name": "Четверг"},
    {"id": 5, "name": "Пятница"},
    {"id": 6, "name": "Суббота"},
)
PAIR_SLOTS = (
    {"number": 1, "label": "1 пара", "time": "9:00-10:30"},
    {"number": 2, "label": "2 пара", "time": "10:40-12:10"},
    {"number": 3, "label": "3 пара", "time": "12:40-14:10"},
    {"number": 4, "label": "4 пара", "time": "14:20-15:50"},
    {"number": 5, "label": "5 пара", "time": "16:20-17:50"},
    {"number": 6, "label": "6 пара", "time": "18:00-19:30"},
    {"number": 7, "label": "7 пара", "time": "19:40-21:10"},
)
PAIR_SLOT_BY_NUMBER = {int(slot["number"]): slot for slot in PAIR_SLOTS}
VALID_DAY_IDS = {int(day["id"]) for day in DAY_OPTIONS}
VALID_PAIR_NUMBERS = {int(slot["number"]) for slot in PAIR_SLOTS}

ATTENDANCE_STATUS_LABELS = {
    ATTENDANCE_STATUS_PRESENT: "Присутствовал",
    ATTENDANCE_STATUS_ABSENT: "Отсутствовал",
    ATTENDANCE_STATUS_EXCUSED: "Отсутствовал (уваж.)",
}
ATTENDANCE_STATUS_SHORT = {
    ATTENDANCE_STATUS_PRESENT: "П",
    ATTENDANCE_STATUS_ABSENT: "Н",
    ATTENDANCE_STATUS_EXCUSED: "У",
}
PUBLIC_ENDPOINTS = {"journal_checkin_page", "static", "favicon"}
MOSCOW_TZ = ZoneInfo("Europe/Moscow") if ZoneInfo is not None else timezone.utc


def register_journal_routes(
    app,
    db,
    Course,
    Group,
    Student,
    JournalLesson,
    JournalLessonSession,
    JournalAttendance,
    parse_int,
):
    runtime = app.extensions.setdefault("journal_runtime", {})
    attendance_events = runtime.get("attendance_events")
    if attendance_events is None:
        attendance_events = RealtimeEventBus()
        runtime["attendance_events"] = attendance_events

    tunnel_events = runtime.get("tunnel_events")
    if tunnel_events is None:
        tunnel_events = RealtimeEventBus()
        runtime["tunnel_events"] = tunnel_events

    tunnel = runtime.get("tunnel")
    if tunnel is None:
        tunnel = JournalTunnelManager()
        runtime["tunnel"] = tunnel
        tunnel.set_on_change(lambda: tunnel_events.bump("tunnel"))

    if not runtime.get("tunnel_atexit_registered"):
        atexit.register(lambda: tunnel.close())
        runtime["tunnel_atexit_registered"] = True

    if "active_public_session_key" not in runtime:
        runtime["active_public_session_key"] = ""

    def _public_session_key(lesson_id: int, lesson_date: date) -> str:
        return f"{int(lesson_id)}:{lesson_date.isoformat()}"

    def _get_active_public_session_key() -> str:
        return str(runtime.get("active_public_session_key") or "")

    def _set_active_public_session_key(value: str) -> None:
        runtime["active_public_session_key"] = str(value or "").strip()

    def _parse_public_session_key(value: str):
        raw = str(value or "").strip()
        if not raw or ":" not in raw:
            return None
        lesson_part, date_part = raw.split(":", 1)
        lesson_id = parse_int(lesson_part, default=0)
        lesson_date = _parse_lesson_date(date_part)
        if lesson_id <= 0 or lesson_date is None:
            return None
        return int(lesson_id), lesson_date

    def _active_public_session_exists(key: str) -> bool:
        parsed = _parse_public_session_key(key)
        if not parsed:
            return False
        lesson_id, lesson_date = parsed
        session = JournalLessonSession.query.filter_by(lesson_id=lesson_id, session_date=lesson_date).first()
        return bool(session and str(session.qr_token or "").strip())

    def _request_host_candidates(req) -> set[str]:
        hosts = set()
        raw_values = [
            str(getattr(req, "host", "") or ""),
            str(req.headers.get("X-Forwarded-Host", "") or ""),
            str(req.headers.get("X-Original-Host", "") or ""),
        ]
        for raw in raw_values:
            for part in str(raw).split(","):
                value = part.strip()
                if not value:
                    continue
                host = value.split(":")[0].strip("[]").lower()
                if host:
                    hosts.add(host)
        return hosts

    def _is_public_tunnel_host(host: str) -> bool:
        safe = str(host or "").strip().lower()
        return safe.endswith(".lhr.life") or safe.endswith(".localhost.run")

    def _in_range(value: date, start_date: date, end_date: date) -> bool:
        return start_date <= value <= end_date

    def _calendar_for_start_year(start_year: int):
        return {
            "autumn_classes_start": date(start_year, 9, 1),
            "autumn_classes_end": date(start_year, 12, 23),
            "autumn_credit_start": date(start_year, 12, 24),
            "autumn_credit_end": date(start_year, 12, 31),
            "new_year_break_start": date(start_year, 12, 31),
            "new_year_break_end": date(start_year + 1, 1, 9),
            "winter_gap_start": date(start_year + 1, 1, 10),
            "winter_gap_end": date(start_year + 1, 1, 11),
            "winter_exam_start": date(start_year + 1, 1, 12),
            "winter_exam_end": date(start_year + 1, 1, 31),
            "winter_holidays_start": date(start_year + 1, 2, 1),
            "winter_holidays_end": date(start_year + 1, 2, 8),
            "spring_classes_start": date(start_year + 1, 2, 9),
            "spring_classes_end": date(start_year + 1, 6, 6),
            "spring_gap_start": date(start_year + 1, 6, 7),
            "spring_gap_end": date(start_year + 1, 6, 10),
            "spring_credit_start": date(start_year + 1, 6, 11),
            "spring_credit_end": date(start_year + 1, 6, 20),
            "summer_exam_start": date(start_year + 1, 6, 21),
            "summer_exam_end": date(start_year + 1, 7, 6),
            "summer_holidays_start": date(start_year + 1, 7, 6),
            "summer_holidays_end": date(start_year + 1, 8, 31),
        }

    def _semester_key(start_year: int, term: int) -> str:
        return f"{start_year}-{start_year + 1}:{term}"

    def _day_id_from_date(value: date) -> int:
        return int(value.isoweekday())

    def _semester_label(semester_key: str) -> str:
        if ":" not in str(semester_key):
            return semester_key
        years, _term = str(semester_key).split(":", 1)
        return str(years)

    def _semester_base_for_date(value: date):
        month = int(value.month)
        year = int(value.year)

        start_year = year if month >= 9 else year - 1
        term = 1 if (month >= 9 or month == 1) else 2
        semester_key = _semester_key(start_year, term)

        return {
            "key": semester_key,
            "label": _semester_label(semester_key),
            "start_year": start_year,
            "term": term,
        }

    def _date_context(value: date):
        semester_base = _semester_base_for_date(value)
        if not semester_base:
            return None

        calendar = _calendar_for_start_year(int(semester_base["start_year"]))
        day_of_week = _day_id_from_date(value)

        ctx = {
            "semester_key": semester_base["key"],
            "semester_label": semester_base["label"],
            "day_of_week": day_of_week,
            "stage": "unknown",
            "week_number": None,
            "week_parity": None,
        }

        stage = "unknown"
        class_start_date = None
        if _in_range(value, calendar["autumn_classes_start"], calendar["autumn_classes_end"]):
            stage = "classes_autumn"
            class_start_date = calendar["autumn_classes_start"]
        elif _in_range(value, calendar["spring_classes_start"], calendar["spring_classes_end"]):
            stage = "classes_spring"
            class_start_date = calendar["spring_classes_start"]
        elif _in_range(value, calendar["new_year_break_start"], calendar["new_year_break_end"]):
            stage = "new_year_break"
        elif _in_range(value, calendar["winter_holidays_start"], calendar["winter_holidays_end"]):
            stage = "winter_holidays"
        elif _in_range(value, calendar["summer_holidays_start"], calendar["summer_holidays_end"]):
            stage = "summer_holidays"
        elif _in_range(value, calendar["autumn_credit_start"], calendar["autumn_credit_end"]):
            stage = "autumn_credit"
        elif _in_range(value, calendar["winter_gap_start"], calendar["winter_gap_end"]):
            stage = "winter_gap"
        elif _in_range(value, calendar["winter_exam_start"], calendar["winter_exam_end"]):
            stage = "winter_exam"
        elif _in_range(value, calendar["spring_gap_start"], calendar["spring_gap_end"]):
            stage = "spring_gap"
        elif _in_range(value, calendar["spring_credit_start"], calendar["spring_credit_end"]):
            stage = "spring_credit"
        elif _in_range(value, calendar["summer_exam_start"], calendar["summer_exam_end"]):
            stage = "summer_exam"

        ctx["stage"] = stage

        if stage in ("classes_autumn", "classes_spring") and class_start_date:
            raw_week = ((value - class_start_date).days // 7) + 1
            week_number = min(max(1, int(raw_week)), 16)
            ctx["week_number"] = week_number
            ctx["week_parity"] = "I" if (week_number % 2 == 1) else "II"

        return ctx

    def _active_semester_base():
        today = date.today()
        current = _semester_base_for_date(today)
        if current:
            return current
        start_year = today.year
        autumn_key = _semester_key(start_year, 1)
        return {"key": autumn_key, "label": _semester_label(autumn_key), "start_year": start_year, "term": 1}

    def _stage_add_error(ctx) -> str:
        stage = str((ctx or {}).get("stage") or "")
        if stage == "autumn_credit":
            return "Идет зачетная сессия (24 декабря - 31 декабря), пары недоступны"
        if stage == "new_year_break":
            return "Идут новогодние выходные (31 декабря - 9 января), пары недоступны"
        if stage == "winter_gap":
            return "Период между праздниками и экзаменационной сессией, пары недоступны"
        if stage == "winter_exam":
            return "Идет зимняя экзаменационная сессия (12 января - 31 января), пары недоступны"
        if stage == "winter_holidays":
            return "Идут зимние каникулы (1 февраля - 8 февраля), пары недоступны"
        if stage == "spring_gap":
            return "Период между занятиями и зачетной сессией, пары недоступны"
        if stage == "spring_credit":
            return "Идет зачетная сессия (11 июня - 20 июня), пары недоступны"
        if stage == "summer_exam":
            return "Идет летняя экзаменационная сессия (21 июня - 6 июля), пары недоступны"
        if stage == "summer_holidays":
            return "Идут летние каникулы (6 июля - 31 августа), пары недоступны"
        return f"Дата вне учебного периода для {ctx['semester_label']}"

    def _cleanup_outdated_lessons(active_semester_key: str) -> None:
        stale_lessons = JournalLesson.query.filter(
            or_(
                JournalLesson.semester_key.is_(None),
                JournalLesson.semester_key == "",
                JournalLesson.semester_key != active_semester_key,
            )
        ).all()
        if not stale_lessons:
            return
        for stale_lesson in stale_lessons:
            db.session.delete(stale_lesson)
        db.session.commit()

    def _parse_lesson_date(raw_value):
        raw = str(raw_value or "").strip()
        if not raw:
            return None
        try:
            return date.fromisoformat(raw)
        except Exception:
            return None

    def _student_count_map():
        rows = db.session.query(Student.group_id, func.count(Student.id)).group_by(Student.group_id).all()
        return {int(group_id): int(count) for group_id, count in rows}

    def _unique_group_ids(values):
        out = []
        for raw in values or []:
            gid = parse_int(raw, default=0)
            if gid > 0 and gid not in out:
                out.append(int(gid))
        return out

    def _parse_int_list(values):
        out = []
        for raw in values or []:
            value = parse_int(raw, default=0)
            if value > 0 and value not in out:
                out.append(int(value))
        return out

    def _normalize_group_ids_csv(values):
        return ",".join(str(gid) for gid in _unique_group_ids(values))

    def _lesson_group_ids(lesson):
        if lesson is None:
            return []
        out = []
        raw_csv = str(getattr(lesson, "group_ids", "") or "").strip()
        if raw_csv:
            out.extend(_unique_group_ids(raw_csv.split(",")))
        primary_group_id = parse_int(getattr(lesson, "group_id", 0), default=0)
        if primary_group_id > 0 and primary_group_id not in out:
            out.insert(0, int(primary_group_id))
        return _unique_group_ids(out)

    def _lesson_primary_group_id(lesson):
        group_ids = _lesson_group_ids(lesson)
        if group_ids:
            return int(group_ids[0])
        return parse_int(getattr(lesson, "group_id", 0), default=0)

    def _groups_map(group_ids):
        ids = _unique_group_ids(group_ids)
        if not ids:
            return {}
        groups = Group.query.filter(Group.id.in_(ids)).all()
        return {int(group.id): group for group in groups}

    def _slot_group_conflicts(semester_key: str, week_parity: str, day_of_week: int, pair_number: int, group_ids, exclude_lesson_id: int = 0):
        requested = set(_unique_group_ids(group_ids))
        if not requested:
            return []
        lessons = JournalLesson.query.filter_by(
            semester_key=str(semester_key),
            week_parity=str(week_parity),
            day_of_week=int(day_of_week),
            pair_number=int(pair_number),
        ).all()
        conflicts = set()
        for item in lessons:
            if exclude_lesson_id and int(item.id) == int(exclude_lesson_id):
                continue
            lesson_groups = set(_lesson_group_ids(item))
            overlaps = requested.intersection(lesson_groups)
            if overlaps:
                conflicts.update(overlaps)
        return sorted(conflicts)

    def _summary_for_session_groups(session_row, group_ids, student_counts):
        ids = _unique_group_ids(group_ids)
        total_students = sum(int(student_counts.get(int(gid), 0)) for gid in ids)
        present_count = 0
        excused_count = 0
        if session_row and ids:
            rows = (
                db.session.query(
                    JournalAttendance.status,
                    func.count(JournalAttendance.id),
                )
                .join(Student, Student.id == JournalAttendance.student_id)
                .filter(
                    JournalAttendance.session_id == session_row.id,
                    Student.group_id.in_(ids),
                )
                .group_by(JournalAttendance.status)
                .all()
            )
            by_status = {str(status): int(count) for status, count in rows}
            present_count = int(by_status.get(ATTENDANCE_STATUS_PRESENT, 0))
            excused_count = int(by_status.get(ATTENDANCE_STATUS_EXCUSED, 0))
        return {
            "total_students": int(total_students),
            "present_count": int(present_count),
            "excused_count": int(excused_count),
            "absent_count": max(int(total_students) - int(present_count) - int(excused_count), 0),
        }

    def _normalize_status(raw_status: str):
        value = str(raw_status or "").strip().lower()
        if value in ATTENDANCE_STATUSES:
            return value
        return None

    def _request_ip() -> str:
        for header in ("X-Forwarded-For", "CF-Connecting-IP", "X-Real-IP"):
            raw = str(request.headers.get(header) or "").strip()
            if raw:
                first = raw.split(",")[0].strip()
                if first:
                    return first[:64]
        return str(request.remote_addr or "")[:64]

    def _as_utc(value):
        if value is None:
            return None
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)

    def _to_moscow(value):
        utc_value = _as_utc(value)
        if utc_value is None:
            return None
        return utc_value.astimezone(MOSCOW_TZ)

    def _format_moscow(value, with_seconds: bool = False) -> str:
        local_value = _to_moscow(value)
        if local_value is None:
            return "-"
        pattern = "%d.%m.%Y %H:%M:%S" if with_seconds else "%d.%m.%Y %H:%M"
        return local_value.strftime(pattern)

    def _build_qr_data_uri(link: str):
        if not link:
            return None, None
        if segno is None:
            return None, "Модуль segno не установлен. Выполните установку: pip install -r requirements.txt"

        try:
            qr = segno.make(link, error="m")
            buffer = io.BytesIO()
            qr.save(buffer, kind="png", scale=7, border=2)
            raw_png = buffer.getvalue()
            encoded = base64.b64encode(raw_png).decode("ascii")
            return f"data:image/png;base64,{encoded}", None
        except Exception as exc:
            return None, f"Не удалось сгенерировать QR: {exc}"

    def _pair_info(pair_number: int):
        return PAIR_SLOT_BY_NUMBER.get(int(pair_number or 0), {"number": pair_number, "label": f"{pair_number} пара", "time": ""})

    def _generate_qr_token():
        return secrets.token_urlsafe(24)

    def _get_or_create_session(lesson, lesson_date: date):
        session_row = JournalLessonSession.query.filter_by(lesson_id=lesson.id, session_date=lesson_date).first()
        if session_row:
            return session_row
        session_row = JournalLessonSession(lesson_id=lesson.id, session_date=lesson_date, qr_token="", qr_token_created_at=None)
        db.session.add(session_row)
        db.session.flush()
        return session_row

    def _ensure_session_token(session_row):
        if str(session_row.qr_token or "").strip():
            return False
        session_row.qr_token = _generate_qr_token()
        session_row.qr_token_created_at = datetime.utcnow()
        return True

    def _validate_lesson_date_for_attendance(lesson, lesson_date: date, active_semester):
        if lesson is None or lesson_date is None:
            return None, "Некорректные параметры занятия"

        active_semester_key = str(active_semester["key"])
        if str(lesson.semester_key) != active_semester_key:
            return None, f"Занятие относится к другому семестру. Активный: {active_semester['label']}"

        lesson_ctx = _date_context(lesson_date)
        if not lesson_ctx:
            return None, "Дата вне учебного периода"

        day_of_week = int(lesson_ctx["day_of_week"])
        if day_of_week == 7:
            return None, "В воскресенье занятия не проводятся"

        if str(lesson_ctx.get("semester_key")) != active_semester_key:
            return None, f"Дата вне активного семестра ({active_semester['label']})"

        if str(lesson_ctx.get("stage")) not in ("classes_autumn", "classes_spring"):
            return None, _stage_add_error(lesson_ctx)

        if int(lesson.day_of_week) != day_of_week:
            return None, "Выбранная дата не совпадает с днем недели этого занятия"

        if str(lesson.week_parity) != str(lesson_ctx.get("week_parity") or ""):
            return None, "Выбранная дата не совпадает с четностью недели этого занятия"

        return lesson_ctx, None

    def _attendance_agg_by_session(session_ids):
        if not session_ids:
            return {}
        rows = (
            db.session.query(
                JournalAttendance.session_id,
                JournalAttendance.status,
                func.count(JournalAttendance.id),
            )
            .filter(JournalAttendance.session_id.in_(session_ids))
            .group_by(JournalAttendance.session_id, JournalAttendance.status)
            .all()
        )
        out = {}
        for session_id, status, count in rows:
            out[(int(session_id), str(status))] = int(count)
        return out

    def _lesson_payload(
        lesson,
        course_titles,
        group_names,
        student_counts,
        present_count=0,
        absent_count=0,
        excused_count=0,
        attendance_url="",
        attendance_date="",
    ):
        group_ids = _lesson_group_ids(lesson)
        primary_group_id = _lesson_primary_group_id(lesson)
        course_id = int(lesson.course_id)
        group_name_list = [group_names.get(int(gid), f"Группа #{gid}") for gid in group_ids]
        if not group_name_list:
            group_name_list = [group_names.get(primary_group_id, f"Группа #{primary_group_id}")]
        student_count = sum(int(student_counts.get(int(gid), 0)) for gid in group_ids)
        return {
            "id": lesson.id,
            "week_parity": lesson.week_parity,
            "day_of_week": lesson.day_of_week,
            "pair_number": lesson.pair_number,
            "semester_key": lesson.semester_key,
            "course_id": lesson.course_id,
            "course_title": course_titles.get(course_id, f"Предмет #{course_id}"),
            "group_id": int(primary_group_id),
            "group_ids": [int(gid) for gid in group_ids],
            "group_names": group_name_list,
            "group_name": ", ".join(group_name_list),
            "room": lesson.room,
            "student_count": student_count,
            "present_count": int(present_count),
            "absent_count": int(absent_count),
            "excused_count": int(excused_count),
            "attendance_url": attendance_url or "",
            "attendance_date": attendance_date or "",
        }

    def _build_lessons_for_date(lesson_date: date, active_semester_key: str):
        lesson_ctx = _date_context(lesson_date)
        if not lesson_ctx:
            return []

        day_of_week = int(lesson_ctx["day_of_week"])
        if day_of_week == 7:
            return []
        if str(lesson_ctx.get("stage")) not in ("classes_autumn", "classes_spring"):
            return []
        if str(lesson_ctx.get("semester_key")) != str(active_semester_key):
            return []

        week_parity = str(lesson_ctx.get("week_parity") or "")
        if week_parity not in WEEK_PARITY_OPTIONS:
            return []

        lessons = (
            JournalLesson.query.filter_by(
                semester_key=active_semester_key,
                week_parity=week_parity,
                day_of_week=day_of_week,
            )
            .order_by(JournalLesson.pair_number.asc(), JournalLesson.id.asc())
            .all()
        )
        if not lessons:
            return []

        lesson_ids = [int(lesson.id) for lesson in lessons]
        course_titles = {int(course_id): title for course_id, title in db.session.query(Course.id, Course.title).all()}
        group_names = {int(group_id): name for group_id, name in db.session.query(Group.id, Group.name).all()}
        student_counts = _student_count_map()

        sessions = JournalLessonSession.query.filter(
            JournalLessonSession.lesson_id.in_(lesson_ids),
            JournalLessonSession.session_date == lesson_date,
        ).all()
        session_by_lesson = {int(session_row.lesson_id): session_row for session_row in sessions}

        payload = []
        for lesson in lessons:
            session_row = session_by_lesson.get(int(lesson.id))
            lesson_group_ids = _lesson_group_ids(lesson)
            if lesson_group_ids:
                summary_totals = _summary_for_session_groups(session_row, lesson_group_ids, student_counts)
                present_count = int(summary_totals["present_count"])
                excused_count = int(summary_totals["excused_count"])
                absent_count = int(summary_totals["absent_count"])
            else:
                present_count = 0
                excused_count = 0
                absent_count = 0

            payload.append(
                _lesson_payload(
                    lesson,
                    course_titles=course_titles,
                    group_names=group_names,
                    student_counts=student_counts,
                    present_count=present_count,
                    absent_count=absent_count,
                    excused_count=excused_count,
                    attendance_url=url_for("journal_lesson_page", lesson_id=lesson.id, date=lesson_date.isoformat()),
                    attendance_date=lesson_date.isoformat(),
                )
            )

        payload.sort(
            key=lambda lesson_payload: (
                int(lesson_payload.get("pair_number") or 0),
                str(lesson_payload.get("group_name") or ""),
            )
        )
        return payload

    def _attendance_rows_for_session(session_row, students):
        student_ids = [int(student.id) for student in students]
        rows = (
            JournalAttendance.query.filter(
                JournalAttendance.session_id == session_row.id,
                JournalAttendance.student_id.in_(student_ids),
            )
            .all()
            if student_ids
            else []
        )
        by_student = {int(row.student_id): row for row in rows}

        out_rows = []
        present_count = 0
        excused_count = 0
        for index, student in enumerate(students, start=1):
            row = by_student.get(int(student.id))
            status = ATTENDANCE_STATUS_ABSENT
            source = ""
            source_ip = ""
            marked_at = None
            if row and row.status in ATTENDANCE_STATUSES:
                status = str(row.status)
                source = str(row.source or "")
                source_ip = str(row.source_ip or "")
                marked_at = row.marked_at

            if status == ATTENDANCE_STATUS_PRESENT:
                present_count += 1
            elif status == ATTENDANCE_STATUS_EXCUSED:
                excused_count += 1

            out_rows.append(
                {
                    "index": index,
                    "id": int(student.id),
                    "fio": student.fio,
                    "status": status,
                    "status_label": ATTENDANCE_STATUS_LABELS.get(status, status),
                    "status_short": ATTENDANCE_STATUS_SHORT.get(status, status),
                    "source": source,
                    "source_ip": source_ip,
                    "marked_at": marked_at,
                    "marked_at_display": _format_moscow(marked_at, with_seconds=True) if marked_at else "-",
                }
            )

        total_students = len(students)
        return out_rows, {
            "total_students": total_students,
            "present_count": present_count,
            "excused_count": excused_count,
            "absent_count": max(total_students - present_count - excused_count, 0),
        }

    def _event_key_date(lesson_date: date) -> str:
        return f"date:{lesson_date.isoformat()}"

    def _event_key_lesson(lesson_id: int, lesson_date: date) -> str:
        return f"lesson:{int(lesson_id)}:{lesson_date.isoformat()}"

    def _bump_date_event(lesson_date: date) -> None:
        attendance_events.bump(_event_key_date(lesson_date))

    def _bump_lesson_event(lesson_id: int, lesson_date: date) -> None:
        attendance_events.bump(_event_key_lesson(lesson_id, lesson_date))

    def _bump_related_events(lesson_id: int, lesson_date: date) -> None:
        _bump_date_event(lesson_date)
        _bump_lesson_event(lesson_id, lesson_date)

    def _is_ajax_request() -> bool:
        requested_with = str(request.headers.get("X-Requested-With") or "").strip().lower()
        if requested_with == "xmlhttprequest":
            return True
        accept = str(request.headers.get("Accept") or "").lower()
        return "application/json" in accept

    def _request_local_port(default: int = 5000) -> int:
        try:
            host_value = str(request.host or "")
            if ":" in host_value:
                port_raw = host_value.rsplit(":", 1)[1]
                return int(port_raw)
        except Exception:
            pass
        return int(default)

    def _session_by_lesson_date(lesson, lesson_date: date):
        return JournalLessonSession.query.filter_by(lesson_id=lesson.id, session_date=lesson_date).first()

    def _default_attendance_rows(students):
        rows = []
        for index, student in enumerate(students, start=1):
            rows.append(
                {
                    "index": index,
                    "id": int(student.id),
                    "fio": student.fio,
                    "status": ATTENDANCE_STATUS_ABSENT,
                    "status_label": ATTENDANCE_STATUS_LABELS[ATTENDANCE_STATUS_ABSENT],
                    "status_short": ATTENDANCE_STATUS_SHORT[ATTENDANCE_STATUS_ABSENT],
                    "source": "",
                    "source_ip": "",
                    "marked_at": None,
                    "marked_at_display": "-",
                }
            )
        total = len(students)
        return rows, {
            "total_students": total,
            "present_count": 0,
            "excused_count": 0,
            "absent_count": total,
        }

    def _serialize_rows_for_api(rows):
        payload = []
        for row in rows:
            marked_at = row.get("marked_at")
            local_marked_at = _to_moscow(marked_at)
            marked_at_display = str(row.get("marked_at_display") or "")
            if not marked_at_display:
                marked_at_display = _format_moscow(marked_at, with_seconds=True) if marked_at else "-"
            payload.append(
                {
                    "id": int(row.get("id") or 0),
                    "index": int(row.get("index") or 0),
                    "fio": str(row.get("fio") or ""),
                    "status": str(row.get("status") or ATTENDANCE_STATUS_ABSENT),
                    "status_label": str(row.get("status_label") or ""),
                    "status_short": str(row.get("status_short") or ""),
                    "source": str(row.get("source") or ""),
                    "source_ip": str(row.get("source_ip") or ""),
                    "marked_at": local_marked_at.isoformat() if local_marked_at else "",
                    "marked_at_display": marked_at_display,
                }
            )
        return payload

    def _recent_qr_marks(session_row, limit: int = 14):
        if session_row is None:
            return []
        marks = (
            db.session.query(
                JournalAttendance.student_id,
                JournalAttendance.marked_at,
                JournalAttendance.source_ip,
                JournalAttendance.source,
                JournalAttendance.status,
                Student.fio,
            )
            .join(Student, Student.id == JournalAttendance.student_id)
            .filter(
                JournalAttendance.session_id == session_row.id,
            )
            .order_by(JournalAttendance.marked_at.desc(), JournalAttendance.id.desc())
            .limit(int(limit))
            .all()
        )

        out = []
        for student_id, marked_at, source_ip, source, status, fio in marks:
            local_marked_at = _to_moscow(marked_at)
            source_value = str(source or "").strip().lower()
            if source_value == "qr":
                source_label = "QR"
            elif source_value == "manual":
                source_label = "Локально"
            else:
                source_label = "Система"
            out.append(
                {
                    "student_id": int(student_id),
                    "fio": str(fio or ""),
                    "source_ip": str(source_ip or ""),
                    "source": source_value,
                    "source_label": source_label,
                    "status": str(status or ""),
                    "status_label": ATTENDANCE_STATUS_LABELS.get(str(status or ""), str(status or "")),
                    "marked_at": local_marked_at.isoformat() if local_marked_at else "",
                    "marked_at_display": _format_moscow(marked_at, with_seconds=True) if marked_at else "-",
                }
            )
        return out

    def _build_checkin_urls(session_row, lesson=None, lesson_date: date | None = None):
        checkin_path = ""
        local_checkin_url = ""
        public_checkin_url = ""
        effective_checkin_url = ""

        if session_row and str(session_row.qr_token or "").strip():
            checkin_path = url_for("journal_checkin_page", token=session_row.qr_token)
            local_checkin_url = url_for("journal_checkin_page", token=session_row.qr_token, _external=True)
            active_key = _get_active_public_session_key()
            current_key = ""
            if lesson is not None and lesson_date is not None:
                current_key = _public_session_key(int(lesson.id), lesson_date)

            if not active_key and current_key:
                snap = tunnel.snapshot()
                if bool(snap.get("active")) and str(snap.get("public_url") or "").strip():
                    _set_active_public_session_key(current_key)
                    active_key = current_key

            if active_key and current_key and active_key == current_key:
                public_checkin_url = tunnel.build_public_url_for_path(checkin_path)
                effective_checkin_url = public_checkin_url

        return {
            "checkin_path": checkin_path,
            "local_checkin_url": local_checkin_url,
            "public_checkin_url": public_checkin_url,
            "effective_checkin_url": effective_checkin_url,
        }

    def _tunnel_payload(lesson=None, lesson_date: date | None = None):
        snap = tunnel.snapshot()
        scoped = lesson is not None and lesson_date is not None
        if scoped:
            active_key = _get_active_public_session_key()
            current_key = _public_session_key(int(lesson.id), lesson_date)
            is_current_session = bool(active_key) and active_key == current_key
        else:
            is_current_session = True

        if not is_current_session:
            return {
                "active": False,
                "public_url": "",
                "error_message": "",
                "reconnecting": False,
                "next_refresh_epoch": None,
                "refresh_interval_seconds": snap.get("refresh_interval_seconds"),
            }

        return {
            "active": bool(snap.get("active", False)),
            "public_url": str(snap.get("public_url") or ""),
            "error_message": str(snap.get("error_message") or ""),
            "reconnecting": bool(snap.get("reconnecting", False)),
            "next_refresh_epoch": snap.get("next_refresh_epoch"),
            "refresh_interval_seconds": snap.get("refresh_interval_seconds"),
        }

    def _lesson_attendance_payload(lesson, lesson_date: date, group_id: int | None = None):
        lesson_group_ids = _lesson_group_ids(lesson)
        active_group_id = parse_int(group_id, default=0)
        if active_group_id not in lesson_group_ids:
            active_group_id = int(lesson_group_ids[0]) if lesson_group_ids else 0

        students = (
            Student.query.filter_by(group_id=active_group_id).order_by(Student.fio.asc()).all()
            if active_group_id > 0
            else []
        )
        session_row = _session_by_lesson_date(lesson, lesson_date)
        if session_row:
            rows, summary = _attendance_rows_for_session(session_row, students)
        else:
            rows, summary = _default_attendance_rows(students)

        student_counts = _student_count_map()
        overall_summary = _summary_for_session_groups(session_row, lesson_group_ids, student_counts)

        return {
            "lesson_id": int(lesson.id),
            "lesson_date": lesson_date.isoformat(),
            "session_id": int(session_row.id) if session_row else None,
            "active_group_id": int(active_group_id) if active_group_id > 0 else None,
            "group_ids": [int(gid) for gid in lesson_group_ids],
            "summary": {
                "total_students": int(summary["total_students"]),
                "present_count": int(summary["present_count"]),
                "absent_count": int(summary["absent_count"]),
                "excused_count": int(summary["excused_count"]),
            },
            "overall_summary": {
                "total_students": int(overall_summary["total_students"]),
                "present_count": int(overall_summary["present_count"]),
                "absent_count": int(overall_summary["absent_count"]),
                "excused_count": int(overall_summary["excused_count"]),
            },
            "students": _serialize_rows_for_api(rows),
            "qr_marks": _recent_qr_marks(session_row),
        }

    def _lesson_qr_payload(lesson, lesson_date: date):
        session_row = _session_by_lesson_date(lesson, lesson_date)
        created = False
        if session_row is None:
            session_row = _get_or_create_session(lesson, lesson_date)
            created = True
        if _ensure_session_token(session_row):
            created = True
        if created:
            db.session.commit()

        checkin_urls = _build_checkin_urls(session_row, lesson=lesson, lesson_date=lesson_date)
        qr_data_uri, qr_error = _build_qr_data_uri(checkin_urls["effective_checkin_url"])
        checkin_summary = _summary_for_session_groups(session_row, _lesson_group_ids(lesson), _student_count_map())
        return {
            "lesson_id": int(lesson.id),
            "lesson_date": lesson_date.isoformat(),
            "session_id": int(session_row.id),
            "local_checkin_url": checkin_urls["local_checkin_url"],
            "public_checkin_url": checkin_urls["public_checkin_url"],
            "effective_checkin_url": checkin_urls["effective_checkin_url"],
            "qr_data_uri": qr_data_uri or "",
            "qr_error": qr_error or "",
            "checkin_summary": {
                "total_students": int(checkin_summary["total_students"]),
                "present_count": int(checkin_summary["present_count"]),
                "absent_count": int(checkin_summary["absent_count"]),
                "excused_count": int(checkin_summary["excused_count"]),
            },
            "tunnel": _tunnel_payload(lesson=lesson, lesson_date=lesson_date),
        }

    def _source_label(source_key: str) -> str:
        source = str(source_key or "").strip().lower()
        if source == "qr":
            return "QR"
        if source == "manual":
            return "Локально"
        if source == "unmarked":
            return "Не отмечено"
        return "-"

    def _normalize_source_filters(values):
        out = []
        for raw in values or []:
            value = str(raw or "").strip().lower()
            if value in {"qr", "manual", "unmarked"} and value not in out:
                out.append(value)
        return out

    def _normalize_status_filters(values):
        out = []
        for raw in values or []:
            value = _normalize_status(raw)
            if value and value not in out:
                out.append(value)
        return out

    def _safe_excel_filename(raw_name: str) -> str:
        safe = re.sub(r'[<>:"/\\|?*\n\r\t]+', "_", str(raw_name or "").strip())
        safe = re.sub(r"\s+", " ", safe).strip(" .")
        return safe or "journal_attendance"

    @app.before_request
    def _restrict_public_routes():
        if is_local_request(request):
            return None
        endpoint = str(request.endpoint or "")
        if endpoint in PUBLIC_ENDPOINTS:
            return None
        return ("Not Found", 404)

    @app.get("/journal")
    def journal_page():
        active_semester = _active_semester_base()
        active_semester_key = active_semester["key"]
        _cleanup_outdated_lessons(active_semester_key)

        selected_date = _parse_lesson_date(request.args.get("date")) or date.today()

        courses = Course.query.filter(Course.archived.is_(False)).order_by(Course.title.asc()).all()
        groups = Group.query.order_by(Group.name.asc()).all()

        lessons = (
            JournalLesson.query.filter_by(semester_key=active_semester_key)
            .order_by(
                case((JournalLesson.week_parity == "I", 0), else_=1),
                JournalLesson.day_of_week.asc(),
                JournalLesson.pair_number.asc(),
                JournalLesson.id.asc(),
            )
            .all()
        )

        course_titles = {int(course_id): title for course_id, title in db.session.query(Course.id, Course.title).all()}
        group_names = {int(group.id): group.name for group in groups}
        student_counts = _student_count_map()

        default_lessons = []
        for lesson in lessons:
            lesson_group_ids = _lesson_group_ids(lesson)
            student_count = sum(int(student_counts.get(int(gid), 0)) for gid in lesson_group_ids)
            default_lessons.append(
                _lesson_payload(
                    lesson,
                    course_titles=course_titles,
                    group_names=group_names,
                    student_counts=student_counts,
                    present_count=0,
                    absent_count=student_count,
                    excused_count=0,
                    attendance_url="",
                    attendance_date="",
                )
            )

        return render_template(
            "journal.html",
            courses=courses,
            groups=groups,
            week_parities=WEEK_PARITY_OPTIONS,
            days=DAY_OPTIONS,
            pair_slots=PAIR_SLOTS,
            active_semester_key=active_semester_key,
            active_semester_label=active_semester["label"],
            selected_date_iso=selected_date.isoformat(),
            lessons=default_lessons,
        )

    @app.get("/api/journal/lessons/<int:lesson_id>/delete-scope-preview")
    def api_journal_delete_lesson_scope_preview(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        scope = str(request.args.get("scope") or "").strip().lower() or "single"
        if scope == "single":
            lessons_to_delete = [lesson]
        elif scope in {"course", "name", "all"}:
            lessons_to_delete = JournalLesson.query.filter_by(
                semester_key=str(lesson.semester_key),
                course_id=int(lesson.course_id),
            ).all()
            scope = "course"
        else:
            return jsonify({"success": False, "error": "Некорректный режим удаления"}), 400

        if not lessons_to_delete:
            return jsonify({"success": False, "error": "Занятия для удаления не найдены"}), 404

        lesson_ids = [int(item.id) for item in lessons_to_delete]
        session_rows = (
            JournalLessonSession.query.filter(JournalLessonSession.lesson_id.in_(lesson_ids))
            .order_by(JournalLessonSession.session_date.asc())
            .all()
        )
        session_ids = [int(row.id) for row in session_rows]
        attendance_count = 0
        if session_ids:
            attendance_count = (
                db.session.query(func.count(JournalAttendance.id))
                .filter(JournalAttendance.session_id.in_(session_ids))
                .scalar()
                or 0
            )

        session_dates = sorted(
            {
                row.session_date.isoformat()
                for row in session_rows
                if row.session_date is not None
            }
        )

        group_ids = []
        for item in lessons_to_delete:
            for gid in _lesson_group_ids(item):
                gid_value = int(gid)
                if gid_value > 0 and gid_value not in group_ids:
                    group_ids.append(gid_value)

        group_map = {}
        if group_ids:
            group_map = {
                int(group.id): group.name
                for group in Group.query.filter(Group.id.in_(group_ids)).all()
            }
        group_names = [group_map.get(int(gid), f"Группа #{gid}") for gid in group_ids]

        course = db.session.get(Course, int(lesson.course_id))
        session_dates_preview = session_dates[:20]
        return jsonify(
            {
                "success": True,
                "scope": scope,
                "course_id": int(lesson.course_id),
                "course_title": course.title if course else f"Предмет #{int(lesson.course_id)}",
                "semester_key": str(lesson.semester_key or ""),
                "lessons_count": int(len(lesson_ids)),
                "sessions_count": int(len(session_rows)),
                "attendance_count": int(attendance_count),
                "group_names": group_names,
                "date_from": session_dates[0] if session_dates else None,
                "date_to": session_dates[-1] if session_dates else None,
                "session_dates_preview": session_dates_preview,
                "session_dates_hidden": int(max(len(session_dates) - len(session_dates_preview), 0)),
            }
        )

    @app.get("/journal/export/attendance")
    def journal_export_attendance_excel():
        date_from = _parse_lesson_date(request.args.get("date_from"))
        date_to = _parse_lesson_date(request.args.get("date_to"))
        if date_from is None or date_to is None:
            return ("Укажите корректный диапазон дат в формате YYYY-MM-DD.", 400)
        if date_from > date_to:
            return ("Дата начала не может быть позже даты окончания.", 400)

        raw_student_query = str(request.args.get("student_query") or "")
        student_query = re.sub(r"\s+", " ", raw_student_query).strip()
        student_query_casefold = student_query.casefold()
        selected_student_id = parse_int(request.args.get("student_id"), default=0)
        selected_student = db.session.get(Student, selected_student_id) if selected_student_id > 0 else None
        if not selected_student:
            selected_student_id = 0
        selected_student_group_id = int(selected_student.group_id) if selected_student else 0

        selected_group_ids = _parse_int_list(request.args.getlist("group_ids"))
        selected_course_ids = _parse_int_list(request.args.getlist("course_ids"))
        selected_statuses = _normalize_status_filters(request.args.getlist("status"))
        selected_sources = _normalize_source_filters(request.args.getlist("source"))

        if not selected_statuses:
            selected_statuses = list(ATTENDANCE_STATUSES)
        if not selected_sources:
            selected_sources = ["qr", "manual", "unmarked"]

        status_filter_set = set(selected_statuses)
        source_filter_set = set(selected_sources)

        all_groups = Group.query.order_by(Group.name.asc()).all()
        group_name_map = {int(group.id): group.name for group in all_groups}
        all_courses = Course.query.order_by(Course.title.asc()).all()
        course_title_map = {int(course.id): course.title for course in all_courses}

        selected_group_ids = [gid for gid in selected_group_ids if int(gid) in group_name_map]
        selected_course_ids = [cid for cid in selected_course_ids if int(cid) in course_title_map]
        if selected_student_group_id > 0 and selected_student_group_id in group_name_map:
            selected_group_ids = [int(selected_student_group_id)]
        selected_group_set = set(selected_group_ids)
        selected_course_set = set(selected_course_ids)

        sessions_query = (
            db.session.query(JournalLessonSession, JournalLesson)
            .join(JournalLesson, JournalLesson.id == JournalLessonSession.lesson_id)
            .filter(
                JournalLessonSession.session_date >= date_from,
                JournalLessonSession.session_date <= date_to,
            )
        )
        if selected_course_set:
            sessions_query = sessions_query.filter(JournalLesson.course_id.in_(selected_course_set))

        session_lesson_rows = (
            sessions_query.order_by(
                JournalLessonSession.session_date.asc(),
                JournalLesson.pair_number.asc(),
                JournalLesson.id.asc(),
            ).all()
        )

        day_name_by_id = {int(item["id"]): str(item["name"]) for item in DAY_OPTIONS}
        students_by_group = {}
        export_rows = []

        for session_row, lesson in session_lesson_rows:
            lesson_group_ids = _lesson_group_ids(lesson)
            if selected_group_set:
                target_group_ids = [gid for gid in lesson_group_ids if gid in selected_group_set]
            else:
                target_group_ids = list(lesson_group_ids)
            if not target_group_ids:
                continue

            student_pairs = []
            for gid in target_group_ids:
                if gid not in students_by_group:
                    students_by_group[gid] = (
                        Student.query.filter_by(group_id=int(gid)).order_by(Student.fio.asc()).all()
                    )
                for student in students_by_group[gid]:
                    if selected_student_id > 0 and int(student.id) != int(selected_student_id):
                        continue
                    student_fio = str(student.fio or "")
                    if selected_student_id <= 0 and student_query_casefold and student_query_casefold not in student_fio.casefold():
                        continue
                    student_pairs.append((int(gid), student))
            if not student_pairs:
                continue

            student_ids = [int(student.id) for _, student in student_pairs]
            attendance_rows = (
                JournalAttendance.query.filter(
                    JournalAttendance.session_id == session_row.id,
                    JournalAttendance.student_id.in_(student_ids),
                ).all()
                if student_ids
                else []
            )
            attendance_by_student = {int(row.student_id): row for row in attendance_rows}

            pair_info = _pair_info(lesson.pair_number)
            lesson_date = session_row.session_date
            lesson_date_iso = lesson_date.isoformat() if lesson_date else "-"
            day_label = day_name_by_id.get(int(lesson.day_of_week), "-")
            pair_label = str(pair_info.get("label") or f"{int(lesson.pair_number)} пара")
            pair_time = str(pair_info.get("time") or "")
            course_title = course_title_map.get(int(lesson.course_id), f"Предмет #{int(lesson.course_id)}")
            room_label = str(lesson.room or "-")

            for group_id, student in student_pairs:
                record = attendance_by_student.get(int(student.id))
                if record is None:
                    status = ATTENDANCE_STATUS_ABSENT
                    source_key = "unmarked"
                    source_ip = "-"
                    marked_at_display = "-"
                else:
                    status = _normalize_status(record.status) or ATTENDANCE_STATUS_ABSENT
                    raw_source = str(record.source or "").strip().lower()
                    source_key = "qr" if raw_source == "qr" else "manual"
                    source_ip = str(record.source_ip or "").strip() or "-"
                    marked_at_display = _format_moscow(record.marked_at, with_seconds=True) if record.marked_at else "-"

                if status not in status_filter_set:
                    continue
                if source_key not in source_filter_set:
                    continue

                status_label = ATTENDANCE_STATUS_LABELS.get(status, status)
                export_rows.append(
                    {
                        "date": lesson_date_iso,
                        "day": day_label,
                        "pair": pair_label,
                        "time": pair_time or "-",
                        "course": course_title,
                        "group": group_name_map.get(int(group_id), f"Группа #{group_id}"),
                        "room": room_label,
                        "student": student.fio,
                        "status": status_label,
                        "presence": "Был" if status == ATTENDANCE_STATUS_PRESENT else "Не был",
                        "source": _source_label(source_key),
                        "ip": source_ip,
                        "marked_at": marked_at_display,
                    }
                )

        wb = Workbook()
        ws = wb.active
        ws.title = "Посещаемость"

        total_columns = 13
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        ws["A1"] = "Выгрузка посещаемости"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        selected_groups_label = (
            ", ".join(group_name_map[int(gid)] for gid in selected_group_ids)
            if selected_group_ids
            else "Все группы"
        )
        selected_courses_label = (
            ", ".join(course_title_map[int(cid)] for cid in selected_course_ids)
            if selected_course_ids
            else "Все предметы"
        )
        status_filter_label = ", ".join(ATTENDANCE_STATUS_LABELS.get(value, value) for value in selected_statuses)
        source_filter_label = ", ".join(_source_label(value) for value in selected_sources)
        if selected_student:
            student_filter_label = f"{selected_student.fio} ({group_name_map.get(selected_student_group_id, f'Группа #{selected_student_group_id}')})"
        elif student_query:
            student_filter_label = student_query
        else:
            student_filter_label = "Все студенты"

        ws["A2"] = f"Сформировано: {_format_moscow(datetime.utcnow(), with_seconds=True)} (МСК)"
        ws["A3"] = f"Период: {date_from.isoformat()} - {date_to.isoformat()}"
        ws["A4"] = f"Группы: {selected_groups_label}"
        ws["A5"] = f"Предметы: {selected_courses_label}"
        ws["A6"] = f"Статусы: {status_filter_label} | Источники: {source_filter_label}"
        ws["A7"] = f"Студент: {student_filter_label}"

        for row_idx in range(2, 8):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_columns)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        header_row = 8
        headers = [
            "Дата",
            "День",
            "Пара",
            "Время",
            "Предмет",
            "Группа",
            "Аудитория",
            "Студент",
            "Статус",
            "Факт",
            "Источник",
            "IP",
            "Отмечено (МСК)",
        ]
        header_fill = PatternFill("solid", fgColor="E9EEF5")
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        current_row = header_row + 1
        if export_rows:
            for item in export_rows:
                ws.cell(row=current_row, column=1, value=item["date"])
                ws.cell(row=current_row, column=2, value=item["day"])
                ws.cell(row=current_row, column=3, value=item["pair"])
                ws.cell(row=current_row, column=4, value=item["time"])
                ws.cell(row=current_row, column=5, value=item["course"])
                ws.cell(row=current_row, column=6, value=item["group"])
                ws.cell(row=current_row, column=7, value=item["room"])
                ws.cell(row=current_row, column=8, value=item["student"])
                ws.cell(row=current_row, column=9, value=item["status"])
                ws.cell(row=current_row, column=10, value=item["presence"])
                ws.cell(row=current_row, column=11, value=item["source"])
                ws.cell(row=current_row, column=12, value=item["ip"])
                ws.cell(row=current_row, column=13, value=item["marked_at"])
                current_row += 1
        else:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_columns)
            ws.cell(row=current_row, column=1, value="По выбранным фильтрам данные не найдены.")
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

        widths = {
            1: 12,
            2: 14,
            3: 11,
            4: 11,
            5: 34,
            6: 18,
            7: 13,
            8: 34,
            9: 21,
            10: 10,
            11: 16,
            12: 18,
            13: 22,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = f"A{header_row + 1}"
        if export_rows:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(total_columns)}{current_row - 1}"

        book_io = io.BytesIO()
        wb.save(book_io)
        book_io.seek(0)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename_raw = (
            f"journal_attendance_{date_from.isoformat()}_{date_to.isoformat()}_{timestamp}"
        )
        filename = f"{_safe_excel_filename(filename_raw)}.xlsx"
        return send_file(
            book_io,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/api/journal/students/search")
    def api_journal_students_search():
        raw_query = str(request.args.get("q") or "")
        query = re.sub(r"\s+", " ", raw_query).strip()
        query_fold = query.casefold()
        if len(query) < 2:
            return jsonify({"success": True, "students": []})

        limit = parse_int(request.args.get("limit"), default=10)
        limit = min(max(limit, 1), 20)

        candidates_query = (
            db.session.query(Student.id, Student.fio, Student.group_id, Group.name)
            .join(Group, Group.id == Student.group_id)
            .order_by(Student.fio.asc())
        )

        sql_like = f"%{query}%"
        candidates = candidates_query.filter(Student.fio.ilike(sql_like)).limit(int(limit * 6)).all()
        if not candidates:
            return jsonify({"success": True, "students": []})

        items = []
        seen_ids = set()
        for student_id, fio, group_id, group_name in candidates:
            sid = int(student_id)
            if sid in seen_ids:
                continue
            safe_fio = str(fio or "").strip()
            if not safe_fio:
                continue
            if query_fold not in safe_fio.casefold():
                continue
            seen_ids.add(sid)
            gid = int(group_id)
            items.append(
                {
                    "id": sid,
                    "fio": safe_fio,
                    "group_id": gid,
                    "group_name": str(group_name or f"Группа #{gid}"),
                }
            )
            if len(items) >= limit:
                break

        return jsonify({"success": True, "students": items})

    @app.get("/api/journal/date/<lesson_date>/lessons")
    def api_journal_lessons_by_date(lesson_date: str):
        lesson_date_value = _parse_lesson_date(lesson_date)
        if lesson_date_value is None:
            return jsonify({"success": False, "error": "Некорректная дата"}), 400

        active_semester = _active_semester_base()
        active_semester_key = str(active_semester["key"])
        _cleanup_outdated_lessons(active_semester_key)

        lessons = _build_lessons_for_date(lesson_date_value, active_semester_key)
        return jsonify({"success": True, "lessons": lessons})

    @app.get("/stream/journal/date/<lesson_date>")
    def stream_journal_date(lesson_date: str):
        lesson_date_value = _parse_lesson_date(lesson_date)
        if lesson_date_value is None:
            return jsonify({"success": False, "error": "Некорректная дата"}), 400

        def _date_payload():
            active_semester = _active_semester_base()
            active_semester_key = str(active_semester["key"])
            lessons = _build_lessons_for_date(lesson_date_value, active_semester_key)
            return {"date": lesson_date_value.isoformat(), "lessons": lessons}

        event_key = _event_key_date(lesson_date_value)

        @stream_with_context
        def generate():
            version = attendance_events.get_version(event_key)
            initial_payload = _date_payload()
            yield f"event: lessons\ndata: {json.dumps(initial_payload, ensure_ascii=False)}\n\n"

            while True:
                next_version = attendance_events.wait_for_change(event_key, version, timeout=30.0)
                if next_version == version:
                    yield ": keepalive\n\n"
                    continue
                version = next_version
                payload = _date_payload()
                yield f"event: lessons\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

        headers = {
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
        return Response(generate(), mimetype="text/event-stream", headers=headers)

    @app.get("/api/journal/group/<int:group_id>/students")
    def api_journal_group_students(group_id: int):
        group = db.session.get(Group, group_id)
        if not group:
            return jsonify({"success": False, "error": "Группа не найдена"}), 404

        students = Student.query.filter_by(group_id=group_id).order_by(Student.fio.asc()).all()
        return jsonify(
            {
                "success": True,
                "group": {"id": group.id, "name": group.name},
                "students": [student.to_dict() for student in students],
            }
        )

    @app.get("/api/journal/groups/students")
    def api_journal_groups_students():
        raw_ids = str(request.args.get("ids") or "").strip()
        if not raw_ids:
            return jsonify({"success": False, "error": "Список групп не передан"}), 400

        parsed_ids = []
        for raw in raw_ids.split(","):
            gid = parse_int(raw, default=0)
            if gid > 0 and gid not in parsed_ids:
                parsed_ids.append(int(gid))
        if not parsed_ids:
            return jsonify({"success": False, "error": "Некорректный список групп"}), 400

        groups = Group.query.filter(Group.id.in_(parsed_ids)).order_by(Group.name.asc()).all()
        by_id = {int(group.id): group for group in groups}
        missing = [gid for gid in parsed_ids if gid not in by_id]
        if missing:
            return jsonify({"success": False, "error": "Одна или несколько групп не найдены"}), 404

        group_payload = []
        total_students = 0
        for gid in parsed_ids:
            group = by_id[gid]
            students = Student.query.filter_by(group_id=gid).order_by(Student.fio.asc()).all()
            total_students += len(students)
            group_payload.append(
                {
                    "group": {"id": int(group.id), "name": group.name},
                    "students": [student.to_dict() for student in students],
                }
            )

        return jsonify(
            {
                "success": True,
                "groups": group_payload,
                "group_count": len(group_payload),
                "total_students": int(total_students),
            }
        )

    @app.post("/api/journal/lessons")
    def api_journal_add_lesson():
        data = request.get_json(silent=True) or {}

        lesson_date = _parse_lesson_date(data.get("date"))
        if lesson_date is None:
            return jsonify({"success": False, "error": "Укажите корректную дату занятия"}), 400

        lesson_ctx = _date_context(lesson_date)
        if not lesson_ctx:
            return jsonify({"success": False, "error": "Не удалось определить параметры даты"}), 400

        day_of_week = int(lesson_ctx["day_of_week"])
        if day_of_week == 7:
            return jsonify({"success": False, "error": "Воскресенье недоступно для добавления пары"}), 400

        active_semester = _active_semester_base()
        active_semester_key = str(active_semester["key"])
        if str(lesson_ctx["semester_key"]) != active_semester_key:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Можно добавлять занятия только для активного семестра: {active_semester['label']}",
                    }
                ),
                400,
            )

        if str(lesson_ctx.get("stage")) not in ("classes_autumn", "classes_spring"):
            return jsonify({"success": False, "error": _stage_add_error(lesson_ctx)}), 400

        _cleanup_outdated_lessons(active_semester_key)

        week_parity = str(lesson_ctx["week_parity"] or "")
        semester_key = str(lesson_ctx["semester_key"])

        pair_number = parse_int(data.get("pair_number"), default=0)
        course_id = parse_int(data.get("course_id"), default=0)
        group_ids = []
        raw_group_ids = data.get("group_ids")
        if isinstance(raw_group_ids, list):
            for raw in raw_group_ids:
                gid = parse_int(raw, default=0)
                if gid > 0 and gid not in group_ids:
                    group_ids.append(int(gid))
        elif isinstance(raw_group_ids, str):
            for raw in raw_group_ids.split(","):
                gid = parse_int(raw, default=0)
                if gid > 0 and gid not in group_ids:
                    group_ids.append(int(gid))

        if not group_ids:
            single_group_id = parse_int(data.get("group_id"), default=0)
            if single_group_id > 0:
                group_ids = [int(single_group_id)]
        room = str(data.get("room") or "").strip()

        if day_of_week not in VALID_DAY_IDS:
            return jsonify({"success": False, "error": "Некорректный день недели"}), 400
        if pair_number not in VALID_PAIR_NUMBERS:
            return jsonify({"success": False, "error": "Некорректный номер пары"}), 400
        if not room:
            return jsonify({"success": False, "error": "Укажите номер аудитории"}), 400

        room = room[:40]

        course = db.session.get(Course, course_id)
        if not course:
            return jsonify({"success": False, "error": "Предмет не найден"}), 404
        if course.archived:
            return jsonify({"success": False, "error": "Нельзя добавить архивный предмет"}), 400

        if not group_ids:
            return jsonify({"success": False, "error": "Выберите хотя бы одну группу"}), 400

        groups = Group.query.filter(Group.id.in_(group_ids)).order_by(Group.name.asc()).all()
        groups_by_id = {int(group.id): group for group in groups}
        missing_groups = [gid for gid in group_ids if gid not in groups_by_id]
        if missing_groups:
            return jsonify({"success": False, "error": "Одна или несколько групп не найдены"}), 404

        conflict_group_ids = _slot_group_conflicts(
            semester_key=semester_key,
            week_parity=week_parity,
            day_of_week=day_of_week,
            pair_number=pair_number,
            group_ids=group_ids,
            exclude_lesson_id=0,
        )
        if conflict_group_ids:
            duplicate_list = ", ".join(
                sorted({groups_by_id[int(gid)].name for gid in conflict_group_ids if int(gid) in groups_by_id})
            )
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Для выбранной пары уже есть занятия у групп: {duplicate_list}",
                    }
                ),
                409,
            )

        lesson = JournalLesson(
            week_parity=week_parity,
            day_of_week=day_of_week,
            pair_number=pair_number,
            semester_key=semester_key,
            course_id=course.id,
            group_id=int(group_ids[0]),
            group_ids=_normalize_group_ids_csv(group_ids),
            room=room,
        )
        db.session.add(lesson)
        db.session.commit()
        _bump_date_event(lesson_date)

        student_counts = _student_count_map()
        group_names_map = {int(group.id): group.name for group in groups}
        student_count = sum(int(student_counts.get(int(gid), 0)) for gid in _lesson_group_ids(lesson))
        payload = _lesson_payload(
            lesson,
            {int(course.id): course.title},
            group_names_map,
            student_counts,
            present_count=0,
            absent_count=student_count,
            excused_count=0,
            attendance_url="",
            attendance_date="",
        )

        return (
            jsonify(
                {
                    "success": True,
                    "lesson": payload,
                    "lessons": [payload],
                    "created_count": 1,
                }
            ),
            201,
        )

    @app.post("/api/journal/lessons/<int:lesson_id>/update")
    def api_journal_update_lesson(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        data = request.get_json(silent=True) or {}
        lesson_date = _parse_lesson_date(
            data.get("date") or request.form.get("date") or request.args.get("date")
        )
        if lesson_date is None:
            return jsonify({"success": False, "error": "Укажите корректную дату занятия"}), 400

        active_semester = _active_semester_base()
        _cleanup_outdated_lessons(str(active_semester["key"]))
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        pair_number = parse_int(data.get("pair_number"), default=0)
        course_id = parse_int(data.get("course_id"), default=0)
        room = str(data.get("room") or "").strip()

        raw_group_ids = data.get("group_ids")
        group_ids = []
        if isinstance(raw_group_ids, list):
            group_ids = _unique_group_ids(raw_group_ids)
        elif isinstance(raw_group_ids, str):
            group_ids = _unique_group_ids(raw_group_ids.split(","))
        if not group_ids:
            fallback_group_id = parse_int(data.get("group_id"), default=0)
            if fallback_group_id > 0:
                group_ids = [int(fallback_group_id)]

        if pair_number not in VALID_PAIR_NUMBERS:
            return jsonify({"success": False, "error": "Некорректный номер пары"}), 400
        if not room:
            return jsonify({"success": False, "error": "Укажите номер аудитории"}), 400
        room = room[:40]

        course = db.session.get(Course, course_id)
        if not course:
            return jsonify({"success": False, "error": "Предмет не найден"}), 404
        if course.archived:
            return jsonify({"success": False, "error": "Нельзя выбрать архивный предмет"}), 400

        if not group_ids:
            return jsonify({"success": False, "error": "Выберите хотя бы одну группу"}), 400

        groups = Group.query.filter(Group.id.in_(group_ids)).order_by(Group.name.asc()).all()
        groups_by_id = {int(group.id): group for group in groups}
        missing_groups = [gid for gid in group_ids if gid not in groups_by_id]
        if missing_groups:
            return jsonify({"success": False, "error": "Одна или несколько групп не найдены"}), 404

        existing_group_ids = _lesson_group_ids(lesson)
        has_history = (
            db.session.query(JournalLessonSession.id)
            .filter(JournalLessonSession.lesson_id == lesson.id)
            .first()
            is not None
        )
        if has_history and not set(existing_group_ids).issubset(set(group_ids)):
            removed_ids = [gid for gid in existing_group_ids if gid not in set(group_ids)]
            removed_groups_map = _groups_map(removed_ids)
            removed_names = ", ".join(
                sorted(
                    {
                        (removed_groups_map.get(int(gid)).name if removed_groups_map.get(int(gid)) else f"Группа #{gid}")
                        for gid in removed_ids
                    }
                )
            )
            return (
                jsonify(
                    {
                        "success": False,
                        "error": (
                            "Для занятия с сохраненной историей нельзя убирать существующие группы. "
                            f"Можно добавить новые. Удаляемые группы: {removed_names}"
                        ),
                    }
                ),
                409,
            )

        conflict_group_ids = _slot_group_conflicts(
            semester_key=str(lesson.semester_key),
            week_parity=str(lesson.week_parity),
            day_of_week=int(lesson.day_of_week),
            pair_number=int(pair_number),
            group_ids=group_ids,
            exclude_lesson_id=int(lesson.id),
        )
        if conflict_group_ids:
            conflict_names = ", ".join(
                sorted({groups_by_id[int(gid)].name for gid in conflict_group_ids if int(gid) in groups_by_id})
            )
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Для выбранной пары уже есть занятия у групп: {conflict_names}",
                    }
                ),
                409,
            )

        lesson.pair_number = int(pair_number)
        lesson.course_id = int(course.id)
        lesson.group_id = int(group_ids[0])
        lesson.group_ids = _normalize_group_ids_csv(group_ids)
        lesson.room = room
        db.session.commit()
        _bump_related_events(lesson.id, lesson_date)

        student_counts = _student_count_map()
        all_group_names = {int(group_id): name for group_id, name in db.session.query(Group.id, Group.name).all()}
        overall_summary = _summary_for_session_groups(
            _session_by_lesson_date(lesson, lesson_date),
            _lesson_group_ids(lesson),
            student_counts,
        )
        payload = _lesson_payload(
            lesson,
            {int(course.id): course.title},
            all_group_names,
            student_counts,
            present_count=overall_summary["present_count"],
            absent_count=overall_summary["absent_count"],
            excused_count=overall_summary["excused_count"],
            attendance_url=url_for("journal_lesson_page", lesson_id=lesson.id, date=lesson_date.isoformat()),
            attendance_date=lesson_date.isoformat(),
        )
        return jsonify({"success": True, "lesson": payload})

    @app.post("/api/journal/lessons/<int:lesson_id>/delete")
    def api_journal_delete_lesson(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        data = request.get_json(silent=True) or {}
        lesson_date = _parse_lesson_date(
            data.get("date") or request.form.get("date") or request.args.get("date")
        )

        db.session.delete(lesson)
        db.session.commit()

        if lesson_date is not None:
            _bump_related_events(lesson_id, lesson_date)

        return jsonify({"success": True})

    @app.post("/api/journal/lessons/<int:lesson_id>/delete-scope")
    def api_journal_delete_lesson_scope(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        data = request.get_json(silent=True) or {}
        lesson_date = _parse_lesson_date(
            data.get("date") or request.form.get("date") or request.args.get("date")
        )
        scope = str(data.get("scope") or "").strip().lower() or "single"

        deleted_ids = []
        deleted_count = 0
        course_id = int(lesson.course_id)

        if scope == "single":
            deleted_ids = [int(lesson.id)]
            db.session.delete(lesson)
            deleted_count = 1
            message = "Занятие удалено"
        elif scope in {"course", "name", "all"}:
            lessons_to_delete = JournalLesson.query.filter_by(
                semester_key=str(lesson.semester_key),
                course_id=int(lesson.course_id),
            ).all()
            if not lessons_to_delete:
                return jsonify({"success": False, "error": "Занятия для удаления не найдены"}), 404
            for item in lessons_to_delete:
                deleted_ids.append(int(item.id))
                db.session.delete(item)
            deleted_count = len(deleted_ids)
            message = f"Удалено занятий по предмету: {deleted_count}"
            scope = "course"
        else:
            return jsonify({"success": False, "error": "Некорректный режим удаления"}), 400

        db.session.commit()

        if lesson_date is not None:
            _bump_date_event(lesson_date)
            for deleted_id in deleted_ids:
                _bump_lesson_event(deleted_id, lesson_date)

            active_key = _get_active_public_session_key()
            for deleted_id in deleted_ids:
                if active_key and active_key == _public_session_key(deleted_id, lesson_date):
                    _set_active_public_session_key("")
                    break

        return jsonify(
            {
                "success": True,
                "scope": scope,
                "deleted_count": int(deleted_count),
                "deleted_ids": [int(item_id) for item_id in deleted_ids],
                "course_id": int(course_id),
                "message": message,
            }
        )

    @app.get("/journal/lesson/<int:lesson_id>")
    def journal_lesson_page(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            flash("Занятие не найдено", "error")
            return redirect(url_for("journal_page"))

        lesson_date = _parse_lesson_date(request.args.get("date"))
        if lesson_date is None:
            flash("Укажите корректную дату занятия", "error")
            return redirect(url_for("journal_page"))

        active_semester = _active_semester_base()
        _cleanup_outdated_lessons(str(active_semester["key"]))

        lesson_ctx, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            flash(validation_error, "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        course = db.session.get(Course, lesson.course_id)
        lesson_group_ids = _lesson_group_ids(lesson)
        lesson_groups_map = _groups_map(lesson_group_ids)
        ordered_group_ids = [gid for gid in lesson_group_ids if gid in lesson_groups_map]

        if course is None or not ordered_group_ids:
            flash("Связанные данные занятия не найдены", "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        selected_group_id = parse_int(request.args.get("group_id"), default=0)
        if selected_group_id not in ordered_group_ids:
            selected_group_id = int(ordered_group_ids[0])
        group = lesson_groups_map.get(int(selected_group_id))
        if group is None:
            flash("Группа занятия не найдена", "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        session_row = _get_or_create_session(lesson, lesson_date)
        changed = _ensure_session_token(session_row)
        if changed:
            db.session.commit()
        else:
            db.session.flush()

        students = Student.query.filter_by(group_id=group.id).order_by(Student.fio.asc()).all()
        student_rows, summary = _attendance_rows_for_session(session_row, students)
        qr_marks = _recent_qr_marks(session_row)

        checkin_urls = _build_checkin_urls(session_row, lesson=lesson, lesson_date=lesson_date)
        qr_data_uri, qr_error = _build_qr_data_uri(checkin_urls["effective_checkin_url"])
        tunnel_state = _tunnel_payload(lesson=lesson, lesson_date=lesson_date)

        group_student_counts = _student_count_map()
        group_name_by_id = {int(gid): lesson_groups_map[int(gid)].name for gid in ordered_group_ids if int(gid) in lesson_groups_map}
        lesson_tabs = []
        for gid in ordered_group_ids:
            lesson_tabs.append(
                {
                    "lesson_id": int(lesson.id),
                    "group_id": int(gid),
                    "group_name": group_name_by_id.get(gid, f"Группа #{gid}"),
                    "student_count": int(group_student_counts.get(gid, 0)),
                    "is_active": int(gid) == int(selected_group_id),
                }
            )

        overall_summary = _summary_for_session_groups(session_row, ordered_group_ids, group_student_counts)
        group_names_display = ", ".join(lesson_groups_map[int(gid)].name for gid in ordered_group_ids if int(gid) in lesson_groups_map)

        pair_info = _pair_info(lesson.pair_number)
        return render_template(
            "journal_lesson.html",
            lesson=lesson,
            lesson_date_iso=lesson_date.isoformat(),
            lesson_ctx=lesson_ctx,
            course=course,
            group=group,
            pair_info=pair_info,
            session_row=session_row,
            summary=summary,
            overall_summary=overall_summary,
            student_rows=student_rows,
            qr_marks=qr_marks,
            status_labels=ATTENDANCE_STATUS_LABELS,
            status_short=ATTENDANCE_STATUS_SHORT,
            local_checkin_url=checkin_urls["local_checkin_url"],
            public_checkin_url=checkin_urls["public_checkin_url"],
            effective_checkin_url=checkin_urls["effective_checkin_url"],
            qr_data_uri=qr_data_uri,
            qr_error=qr_error,
            tunnel=tunnel_state,
            active_semester_label=active_semester["label"],
            lesson_tabs=lesson_tabs,
            active_group_id=int(selected_group_id),
            group_names_display=group_names_display,
        )

    @app.get("/journal/lesson/<int:lesson_id>/qr/view")
    def journal_lesson_qr_view_page(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            flash("Занятие не найдено", "error")
            return redirect(url_for("journal_page"))

        lesson_date = _parse_lesson_date(request.args.get("date"))
        if lesson_date is None:
            flash("Укажите корректную дату занятия", "error")
            return redirect(url_for("journal_page"))

        active_semester = _active_semester_base()
        _cleanup_outdated_lessons(str(active_semester["key"]))
        lesson_ctx, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            flash(validation_error, "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        course = db.session.get(Course, lesson.course_id)
        lesson_group_ids = _lesson_group_ids(lesson)
        lesson_groups_map = _groups_map(lesson_group_ids)
        ordered_group_ids = [gid for gid in lesson_group_ids if gid in lesson_groups_map]
        if course is None or not ordered_group_ids:
            flash("Связанные данные занятия не найдены", "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        active_group_id = parse_int(request.args.get("group_id"), default=0)
        if active_group_id not in ordered_group_ids:
            active_group_id = int(ordered_group_ids[0])

        qr_payload = _lesson_qr_payload(lesson, lesson_date)
        group_names_display = ", ".join(
            lesson_groups_map[int(gid)].name for gid in ordered_group_ids if int(gid) in lesson_groups_map
        )
        pair_info = _pair_info(lesson.pair_number)
        return render_template(
            "journal_qr_view.html",
            lesson=lesson,
            lesson_date_iso=lesson_date.isoformat(),
            lesson_ctx=lesson_ctx,
            course=course,
            pair_info=pair_info,
            qr=qr_payload,
            group_names_display=group_names_display,
            active_group_id=int(active_group_id),
            active_semester_label=active_semester["label"],
        )

    @app.get("/api/journal/lesson/<int:lesson_id>/attendance")
    def api_journal_lesson_attendance(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        lesson_date = _parse_lesson_date(request.args.get("date"))
        if lesson_date is None:
            return jsonify({"success": False, "error": "Некорректная дата занятия"}), 400
        group_id = parse_int(request.args.get("group_id"), default=0)

        active_semester = _active_semester_base()
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        return jsonify({"success": True, "attendance": _lesson_attendance_payload(lesson, lesson_date, group_id=group_id)})

    @app.get("/api/journal/lesson/<int:lesson_id>/qr")
    def api_journal_lesson_qr(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        lesson_date = _parse_lesson_date(request.args.get("date"))
        if lesson_date is None:
            return jsonify({"success": False, "error": "Некорректная дата занятия"}), 400

        active_semester = _active_semester_base()
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        return jsonify({"success": True, "qr": _lesson_qr_payload(lesson, lesson_date)})

    @app.get("/stream/journal/lesson/<int:lesson_id>/<lesson_date>/attendance")
    def stream_journal_lesson_attendance(lesson_id: int, lesson_date: str):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            return jsonify({"success": False, "error": "Занятие не найдено"}), 404

        lesson_date_value = _parse_lesson_date(lesson_date)
        if lesson_date_value is None:
            return jsonify({"success": False, "error": "Некорректная дата занятия"}), 400
        group_id = parse_int(request.args.get("group_id"), default=0)

        active_semester = _active_semester_base()
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date_value, active_semester)
        if validation_error:
            return jsonify({"success": False, "error": validation_error}), 400

        event_key = _event_key_lesson(lesson_id, lesson_date_value)

        @stream_with_context
        def generate():
            version = attendance_events.get_version(event_key)
            initial_payload = _lesson_attendance_payload(lesson, lesson_date_value, group_id=group_id)
            yield f"event: attendance\ndata: {json.dumps(initial_payload, ensure_ascii=False)}\n\n"

            while True:
                next_version = attendance_events.wait_for_change(event_key, version, timeout=30.0)
                if next_version == version:
                    yield ": keepalive\n\n"
                    continue
                version = next_version
                payload = _lesson_attendance_payload(lesson, lesson_date_value, group_id=group_id)
                yield f"event: attendance\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

        headers = {
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
        return Response(generate(), mimetype="text/event-stream", headers=headers)

    @app.get("/stream/journal/tunnel")
    def stream_journal_tunnel():
        event_key = "tunnel"
        scoped_lesson = None
        scoped_lesson_date = _parse_lesson_date(request.args.get("date"))
        scoped_lesson_id = parse_int(request.args.get("lesson_id"), default=0)
        if scoped_lesson_id > 0 and scoped_lesson_date is not None:
            scoped_lesson = db.session.get(JournalLesson, scoped_lesson_id)

        @stream_with_context
        def generate():
            version = tunnel_events.get_version(event_key)
            initial_payload = _tunnel_payload(lesson=scoped_lesson, lesson_date=scoped_lesson_date)
            initial_payload["version"] = int(version)
            yield f"event: state\ndata: {json.dumps(initial_payload, ensure_ascii=False)}\n\n"

            while True:
                next_version = tunnel_events.wait_for_change(event_key, version, timeout=30.0)
                if next_version == version:
                    yield ": keepalive\n\n"
                    continue
                version = next_version
                payload = _tunnel_payload(lesson=scoped_lesson, lesson_date=scoped_lesson_date)
                payload["version"] = int(version)
                yield f"event: state\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n"

        headers = {
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        }
        return Response(generate(), mimetype="text/event-stream", headers=headers)

    @app.post("/api/journal/qr/open")
    @app.post("/journal/qr/open")
    def journal_open_public_qr():
        if not is_local_request(request):
            return jsonify({"success": False, "error": "Открытие публичного QR доступно только локально"}), 403

        data = request.get_json(silent=True) or {}
        lesson_id = parse_int(data.get("lesson_id") or request.form.get("lesson_id"), default=0)
        lesson_date = _parse_lesson_date(data.get("date") or request.form.get("date"))

        lesson = db.session.get(JournalLesson, lesson_id) if lesson_id > 0 else None
        if not lesson or lesson_date is None:
            error_text = "Некорректные параметры занятия"
            if _is_ajax_request():
                return jsonify({"success": False, "error": error_text}), 400
            flash(error_text, "error")
            return redirect(url_for("journal_page"))

        active_semester = _active_semester_base()
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            if _is_ajax_request():
                return jsonify({"success": False, "error": validation_error}), 400
            flash(validation_error, "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        intended_key = _public_session_key(lesson.id, lesson_date)
        _set_active_public_session_key(intended_key)
        local_port = _request_local_port(default=5000)
        ok, message = tunnel.open(local_port=local_port, local_host="127.0.0.1")
        snap = tunnel.snapshot()
        if not ok and bool(snap.get("active")):
            ok = True
            message = "Публичный туннель запускается. Подождите несколько секунд."
        if ok:
            _bump_lesson_event(lesson.id, lesson_date)
            tunnel_events.bump("tunnel")
        else:
            if not bool(snap.get("active")):
                _set_active_public_session_key("")

        if _is_ajax_request():
            payload = _lesson_qr_payload(lesson, lesson_date) if ok else {}
            response_payload = {"success": bool(ok), "message": message, "qr": payload}
            if not ok:
                response_payload["error"] = message
            return jsonify(response_payload)

        flash(message, "success" if ok else "error")
        return redirect(url_for("journal_lesson_page", lesson_id=lesson.id, date=lesson_date.isoformat()))

    @app.post("/api/journal/qr/close")
    @app.post("/journal/qr/close")
    def journal_close_public_qr():
        if not is_local_request(request):
            return jsonify({"success": False, "error": "Закрытие публичного QR доступно только локально"}), 403

        data = request.get_json(silent=True) or {}
        lesson_id = parse_int(data.get("lesson_id") or request.form.get("lesson_id"), default=0)
        lesson_date = _parse_lesson_date(data.get("date") or request.form.get("date"))

        ok, message = tunnel.close(manual=True)
        _set_active_public_session_key("")
        tunnel_events.bump("tunnel")

        if lesson_id > 0 and lesson_date is not None:
            _bump_lesson_event(lesson_id, lesson_date)

        if _is_ajax_request():
            response_payload = {"success": bool(ok), "message": message}
            if not ok:
                response_payload["error"] = message
            return jsonify(response_payload)

        if lesson_id > 0 and lesson_date is not None:
            flash(message, "success" if ok else "error")
            return redirect(url_for("journal_lesson_page", lesson_id=lesson_id, date=lesson_date.isoformat()))
        flash(message, "success" if ok else "error")
        return redirect(url_for("journal_page"))

    @app.post("/journal/lesson/<int:lesson_id>/attendance")
    def journal_set_attendance(lesson_id: int):
        ajax = _is_ajax_request()
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            if ajax:
                return jsonify({"success": False, "error": "Занятие не найдено"}), 404
            flash("Занятие не найдено", "error")
            return redirect(url_for("journal_page"))

        lesson_date = _parse_lesson_date(request.form.get("date"))
        if lesson_date is None:
            if ajax:
                return jsonify({"success": False, "error": "Некорректная дата занятия"}), 400
            flash("Некорректная дата занятия", "error")
            return redirect(url_for("journal_lesson_page", lesson_id=lesson_id))

        active_semester = _active_semester_base()
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            if ajax:
                return jsonify({"success": False, "error": validation_error}), 400
            flash(validation_error, "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        lesson_group_ids = set(_lesson_group_ids(lesson))
        active_group_id = parse_int(request.form.get("group_id"), default=0)
        if active_group_id not in lesson_group_ids:
            active_group_id = 0

        student_id = parse_int(request.form.get("student_id"), default=0)
        student = db.session.get(Student, student_id)
        if not student or int(student.group_id) not in lesson_group_ids:
            if ajax:
                return jsonify({"success": False, "error": "Студент не найден в группе занятия"}), 404
            flash("Студент не найден в группе занятия", "error")
            redirect_kwargs = {"lesson_id": lesson_id, "date": lesson_date.isoformat()}
            if active_group_id > 0:
                redirect_kwargs["group_id"] = active_group_id
            return redirect(url_for("journal_lesson_page", **redirect_kwargs))

        status = _normalize_status(request.form.get("status"))
        if status is None:
            if ajax:
                return jsonify({"success": False, "error": "Некорректный статус посещаемости"}), 400
            flash("Некорректный статус посещаемости", "error")
            redirect_kwargs = {"lesson_id": lesson_id, "date": lesson_date.isoformat()}
            if active_group_id > 0:
                redirect_kwargs["group_id"] = active_group_id
            return redirect(url_for("journal_lesson_page", **redirect_kwargs))

        session_row = _get_or_create_session(lesson, lesson_date)

        record = JournalAttendance.query.filter_by(session_id=session_row.id, student_id=student.id).first()
        now_value = datetime.utcnow()
        source_ip = _request_ip()

        if record:
            record.status = status
            record.source = "manual"
            record.source_ip = source_ip
            record.marked_at = now_value
        else:
            db.session.add(
                JournalAttendance(
                    session_id=session_row.id,
                    student_id=student.id,
                    status=status,
                    source="manual",
                    source_ip=source_ip,
                    marked_at=now_value,
                )
            )

        db.session.commit()
        _bump_related_events(lesson_id, lesson_date)

        if ajax:
            payload = _lesson_attendance_payload(lesson, lesson_date, group_id=active_group_id)
            return jsonify({"success": True, "attendance": payload})

        flash(f"{student.fio}: {ATTENDANCE_STATUS_LABELS.get(status, status)}", "success")
        redirect_kwargs = {"lesson_id": lesson_id, "date": lesson_date.isoformat()}
        if active_group_id > 0:
            redirect_kwargs["group_id"] = active_group_id
        return redirect(url_for("journal_lesson_page", **redirect_kwargs))

    @app.post("/journal/lesson/<int:lesson_id>/qr/regenerate")
    def journal_regenerate_qr(lesson_id: int):
        lesson = db.session.get(JournalLesson, lesson_id)
        if not lesson:
            flash("Занятие не найдено", "error")
            return redirect(url_for("journal_page"))

        lesson_date = _parse_lesson_date(request.form.get("date"))
        if lesson_date is None:
            flash("Некорректная дата занятия", "error")
            return redirect(url_for("journal_lesson_page", lesson_id=lesson_id))

        active_semester = _active_semester_base()
        _, validation_error = _validate_lesson_date_for_attendance(lesson, lesson_date, active_semester)
        if validation_error:
            flash(validation_error, "error")
            return redirect(url_for("journal_page", date=lesson_date.isoformat()))

        session_row = _get_or_create_session(lesson, lesson_date)
        session_row.qr_token = _generate_qr_token()
        session_row.qr_token_created_at = datetime.utcnow()
        db.session.commit()
        _bump_lesson_event(lesson_id, lesson_date)

        if _is_ajax_request():
            return jsonify({"success": True, "qr": _lesson_qr_payload(lesson, lesson_date)})
        flash("QR-ссылка обновлена", "success")
        return redirect(url_for("journal_lesson_page", lesson_id=lesson_id, date=lesson_date.isoformat()))

    @app.route("/journal/checkin/<string:token>", methods=["GET", "POST"])
    def journal_checkin_page(token: str):
        safe_token = str(token or "").strip()
        session_row = (
            JournalLessonSession.query.filter(JournalLessonSession.qr_token == safe_token).order_by(JournalLessonSession.id.desc()).first()
            if safe_token
            else None
        )

        access_error = None
        if session_row is None:
            access_error = "Недействительная или устаревшая QR-ссылка. Попросите преподавателя обновить QR."

        lesson = db.session.get(JournalLesson, int(session_row.lesson_id)) if session_row else None
        if session_row and lesson is None:
            access_error = "Занятие для этой QR-ссылки не найдено."

        if session_row and lesson and not is_local_request(request):
            active_public_key = _get_active_public_session_key()
            current_key = _public_session_key(lesson.id, session_row.session_date)

            snap = tunnel.snapshot()
            snap_public_url = str(snap.get("public_url") or "").strip()
            snap_public_host = (urlparse(snap_public_url).hostname or "").lower() if snap_public_url else ""
            request_hosts = _request_host_candidates(request)
            tunnel_is_active = bool(snap.get("active")) and bool(snap_public_url)
            request_via_active_tunnel = tunnel_is_active and snap_public_host and snap_public_host in request_hosts
            request_via_tunnel_host = any(_is_public_tunnel_host(host) for host in request_hosts)
            can_recover_active_key = bool(
                tunnel_is_active
                and (
                    request_via_active_tunnel
                    or request_via_tunnel_host
                    or not active_public_key
                    or not _active_public_session_exists(active_public_key)
                )
            )

            if can_recover_active_key and (not active_public_key or not _active_public_session_exists(active_public_key)):
                _set_active_public_session_key(current_key)
                active_public_key = current_key

            if not active_public_key or active_public_key != current_key:
                access_error = "Эта QR-ссылка сейчас неактивна. Попросите преподавателя открыть QR для текущего занятия."

        course = db.session.get(Course, int(lesson.course_id)) if lesson else None
        lesson_group_ids = _lesson_group_ids(lesson) if lesson else []
        lesson_groups_map = _groups_map(lesson_group_ids) if lesson_group_ids else {}
        ordered_group_ids = [gid for gid in lesson_group_ids if gid in lesson_groups_map]
        allowed_group_ids = set(ordered_group_ids)
        students = (
            Student.query.filter(Student.group_id.in_(ordered_group_ids)).order_by(Student.fio.asc()).all()
            if ordered_group_ids
            else []
        )
        student_options = []
        for student in students:
            gid = int(student.group_id)
            group_obj = lesson_groups_map.get(gid)
            student_options.append(
                {
                    "id": int(student.id),
                    "fio": student.fio,
                    "group_id": gid,
                    "group_name": group_obj.name if group_obj is not None else f"Группа #{gid}",
                }
            )
        group_names_display = ", ".join(
            lesson_groups_map[int(gid)].name for gid in ordered_group_ids if int(gid) in lesson_groups_map
        )
        has_multiple_groups = len(ordered_group_ids) > 1

        selected_student_id = None
        done_message = None
        done_type = "success"

        if request.method == "POST" and not access_error and lesson and session_row and ordered_group_ids:
            student_id = parse_int(request.form.get("student_id"), default=0)
            selected_student_id = student_id
            student = db.session.get(Student, student_id)
            if not student or int(student.group_id) not in allowed_group_ids:
                done_message = "Выберите себя из списка группы."
                done_type = "error"
            else:
                attendance = JournalAttendance.query.filter_by(session_id=session_row.id, student_id=student.id).first()
                if attendance and attendance.status == ATTENDANCE_STATUS_PRESENT:
                    done_message = f"{student.fio}, вы уже отмечены."
                    done_type = "info"
                else:
                    now_value = datetime.utcnow()
                    source_ip = _request_ip()
                    if attendance:
                        attendance.status = ATTENDANCE_STATUS_PRESENT
                        attendance.source = "qr"
                        attendance.source_ip = source_ip
                        attendance.marked_at = now_value
                    else:
                        db.session.add(
                            JournalAttendance(
                                session_id=session_row.id,
                                student_id=student.id,
                                status=ATTENDANCE_STATUS_PRESENT,
                                source="qr",
                                source_ip=source_ip,
                                marked_at=now_value,
                            )
                        )
                    db.session.commit()
                    _bump_related_events(lesson.id, session_row.session_date)
                    done_message = f"{student.fio}, отметка сохранена."
                    done_type = "success"

        lesson_date_iso = session_row.session_date.isoformat() if session_row and session_row.session_date else ""
        pair_info = _pair_info(lesson.pair_number if lesson else 0)

        return render_template(
            "journal_checkin.html",
            access_error=access_error,
            token=safe_token,
            session_row=session_row,
            lesson=lesson,
            course=course,
            students=student_options,
            selected_student_id=selected_student_id,
            done_message=done_message,
            done_type=done_type,
            lesson_date_iso=lesson_date_iso,
            pair_info=pair_info,
            group_names_display=group_names_display,
            has_multiple_groups=has_multiple_groups,
        )
