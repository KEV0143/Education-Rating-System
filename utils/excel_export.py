import re
from io import BytesIO
from datetime import datetime

from flask import send_file, abort
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _safe_sheet_title(name: str, used: set[str]) -> str:
    name = (name or "").strip() or "Группа"
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    name = name[:31].strip() or "Группа"

    base = name
    i = 2
    while name in used:
        suffix = f" ({i})"
        name = (base[: (31 - len(suffix))] + suffix).strip()
        i += 1
    used.add(name)
    return name


def _style_headers(ws, last_col: int):
    header_fill = PatternFill("solid", fgColor="F1F3F5")
    bold = Font(bold=True)

    for r in (2, 3):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 22


def _auto_fit_some(ws, widths: dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def register_excel_export_routes(app, db, Course, Group, Student, Practice, PracticeGrade, parse_group_ids):

    @app.get("/course/<int:course_id>/export_excel")
    def course_export_excel(course_id: int):
        course = db.session.get(Course, course_id)
        if not course:
            abort(404)

        practice_list = (
            Practice.query.filter_by(course_id=course.id)
            .order_by(Practice.id.asc())
            .all()
        )

        group_ids = parse_group_ids(course.group_ids)
        groups = []
        if group_ids:
            groups = Group.query.filter(Group.id.in_(group_ids)).order_by(Group.name).all()

        wb = Workbook()
        wb.remove(wb.active)

        used_titles = set()
        center = Alignment(horizontal="center", vertical="center")
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fio_align = Alignment(vertical="center")

        if not groups:
            ws = wb.create_sheet(_safe_sheet_title("Нет групп", used_titles))
            ws["A1"] = f"В курсе «{course.title}» нет групп для выгрузки."
            ws["A1"].font = Font(bold=True)
        else:
            for g in groups:
                title = _safe_sheet_title(g.name, used_titles)
                ws = wb.create_sheet(title)

                ws["A1"] = f"{course.title} · {course.year} · {course.semester} семестр · Группа: {g.name}"
                ws["A1"].font = Font(bold=True, size=12)
                ws["A1"].alignment = Alignment(vertical="center")
                ws.cell(row=2, column=1, value="№")
                ws.cell(row=2, column=2, value="ФИО")
                ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
                ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
                col = 3
                for p in practice_list:
                    ws.cell(row=2, column=col, value=p.title)
                    ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
                    ws.cell(row=3, column=col, value="Балл")
                    ws.cell(row=3, column=col + 1, value="Комментарий")
                    col += 2

                last_col = col - 1
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, last_col))

                ws.freeze_panes = "C4"
                ws.auto_filter.ref = f"A3:{get_column_letter(last_col)}3"

                students = (
                    Student.query.filter_by(group_id=g.id)
                    .order_by(Student.fio.asc())
                    .all()
                )

                pids = [p.id for p in practice_list]
                grade_map = {}
                if students and pids:
                    student_ids = [s.id for s in students]
                    rows = (
                        db.session.query(
                            PracticeGrade.student_id,
                            PracticeGrade.practice_id,
                            PracticeGrade.score,
                            PracticeGrade.comment
                        )
                        .filter(
                            PracticeGrade.student_id.in_(student_ids),
                            PracticeGrade.practice_id.in_(pids)
                        )
                        .all()
                    )
                    for sid, pid, score, comment in rows:
                        grade_map[(sid, pid)] = (score, comment or "")

                r = 4
                for idx, s in enumerate(students, start=1):
                    c1 = ws.cell(row=r, column=1, value=idx)
                    c1.alignment = center

                    c2 = ws.cell(row=r, column=2, value=s.fio)
                    c2.alignment = fio_align

                    col = 3
                    for p in practice_list:
                        score, comment = grade_map.get((s.id, p.id), (None, ""))

                        sc_cell = ws.cell(row=r, column=col, value=score)
                        sc_cell.alignment = center

                        cm_cell = ws.cell(row=r, column=col + 1, value=comment)
                        cm_cell.alignment = center_wrap

                        col += 2

                    r += 1

                _style_headers(ws, last_col)

                widths = {1: 5, 2: 34}
                c = 3
                for _ in practice_list:
                    widths[c] = 10
                    widths[c + 1] = 32
                    c += 2
                _auto_fit_some(ws, widths)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"{course.title}_({course.year})_sem{course.semester}_{ts}.xlsx"
        filename = re.sub(r'[<>:"/\\|?*\n\r\t]', "_", filename)

        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
