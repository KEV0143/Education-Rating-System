import io
import re
from datetime import datetime, timezone
from flask import request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.journal.helpers import (
    DAY_OPTIONS,
    ATTENDANCE_STATUSES,
    ATTENDANCE_STATUS_LABELS,
    ATTENDANCE_STATUS_ABSENT,
    ATTENDANCE_STATUS_PRESENT,
    _parse_lesson_date,
    _parse_int_list,
    _normalize_status_filters,
    _normalize_source_filters,
    _lesson_group_ids,
    _pair_info,
    _normalize_status,
    _format_moscow,
    _source_label,
    _safe_excel_filename,
    MOSCOW_TZ
)


def register_journal_export_routes(
    app,
    db,
    Course,
    Group,
    Student,
    JournalLesson,
    JournalLessonSession,
    JournalAttendance,
    parse_int,
):
    @app.get("/journal/export/attendance-excel")
    def journal_export_attendance_excel():
        date_from = _parse_lesson_date(request.args.get("date_from"))
        date_to = _parse_lesson_date(request.args.get("date_to"))
        if date_from is None or date_to is None:
            return ("Укажите корректный диапазон дат в формате YYYY-MM-DD.", 400)
        if date_from > date_to:
            return ("Дата начала не может быть позже даты окончания.", 400)

        raw_student_query = str(request.args.get("student_query") or "")
        student_query = re.sub(r"\s+", " ", raw_student_query).strip()
        student_query_casefold = student_query.casefold()
        selected_student_id = parse_int(request.args.get("student_id"), default=0)
        selected_student = db.session.get(Student, selected_student_id) if selected_student_id > 0 else None
        if not selected_student:
            selected_student_id = 0
        selected_student_group_id = int(selected_student.group_id) if selected_student else 0

        selected_group_ids = _parse_int_list(request.args.getlist("group_ids"))
        selected_course_ids = _parse_int_list(request.args.getlist("course_ids"))
        selected_statuses = _normalize_status_filters(request.args.getlist("status"))
        selected_sources = _normalize_source_filters(request.args.getlist("source"))

        if not selected_statuses:
            selected_statuses = list(ATTENDANCE_STATUSES)
        if not selected_sources:
            selected_sources = ["qr", "manual", "unmarked"]

        status_filter_set = set(selected_statuses)
        source_filter_set = set(selected_sources)

        all_groups = Group.query.order_by(Group.name.asc()).all()
        group_name_map = {int(group.id): group.name for group in all_groups}
        all_courses = Course.query.order_by(Course.title.asc()).all()
        course_title_map = {int(course.id): course.title for course in all_courses}

        selected_group_ids = [gid for gid in selected_group_ids if int(gid) in group_name_map]
        selected_course_ids = [cid for cid in selected_course_ids if int(cid) in course_title_map]
        if selected_student_group_id > 0 and selected_student_group_id in group_name_map:
            selected_group_ids = [int(selected_student_group_id)]
        selected_group_set = set(selected_group_ids)
        selected_course_set = set(selected_course_ids)

        sessions_query = (
            db.session.query(JournalLessonSession, JournalLesson)
            .join(JournalLesson, JournalLesson.id == JournalLessonSession.lesson_id)
            .filter(
                JournalLessonSession.session_date >= date_from,
                JournalLessonSession.session_date <= date_to,
            )
        )
        if selected_course_set:
            sessions_query = sessions_query.filter(JournalLesson.course_id.in_(selected_course_set))

        session_lesson_rows = (
            sessions_query.order_by(
                JournalLessonSession.session_date.asc(),
                JournalLesson.pair_number.asc(),
                JournalLesson.id.asc(),
            ).all()
        )

        day_name_by_id = {int(item["id"]): str(item["name"]) for item in DAY_OPTIONS}
        students_by_group = {}
        export_rows = []

        for session_row, lesson in session_lesson_rows:
            lesson_group_ids = _lesson_group_ids(lesson)
            if selected_group_set:
                target_group_ids = [gid for gid in lesson_group_ids if gid in selected_group_set]
            else:
                target_group_ids = list(lesson_group_ids)
            if not target_group_ids:
                continue

            student_pairs = []
            for gid in target_group_ids:
                if gid not in students_by_group:
                    students_by_group[gid] = (
                        Student.query.filter_by(group_id=int(gid)).order_by(Student.fio.asc()).all()
                    )
                for student in students_by_group[gid]:
                    if selected_student_id > 0 and int(student.id) != int(selected_student_id):
                        continue
                    student_fio = str(student.fio or "")
                    if selected_student_id <= 0 and student_query_casefold and student_query_casefold not in student_fio.casefold():
                        continue
                    student_pairs.append((int(gid), student))
            if not student_pairs:
                continue

            student_ids = [int(student.id) for _, student in student_pairs]
            attendance_rows = (
                JournalAttendance.query.filter(
                    JournalAttendance.session_id == session_row.id,
                    JournalAttendance.student_id.in_(student_ids),
                ).all()
                if student_ids
                else []
            )
            attendance_by_student = {int(row.student_id): row for row in attendance_rows}

            pair_info = _pair_info(lesson.pair_number)
            lesson_date = session_row.session_date
            lesson_date_iso = lesson_date.isoformat() if lesson_date else "-"
            day_label = day_name_by_id.get(int(lesson.day_of_week), "-")
            pair_label = str(pair_info.get("label") or f"{int(lesson.pair_number)} пара")
            pair_time = str(pair_info.get("time") or "")
            course_title = course_title_map.get(int(lesson.course_id), f"Предмет #{int(lesson.course_id)}")
            room_label = str(lesson.room or "-")

            for group_id, student in student_pairs:
                record = attendance_by_student.get(int(student.id))
                if record is None:
                    status = ATTENDANCE_STATUS_ABSENT
                    source_key = "unmarked"
                    source_ip = "-"
                    marked_at_display = "-"
                else:
                    status = _normalize_status(record.status) or ATTENDANCE_STATUS_ABSENT
                    raw_source = str(record.source or "").strip().lower()
                    source_key = "qr" if raw_source == "qr" else "manual"
                    source_ip = str(record.source_ip or "").strip() or "-"
                    marked_at_display = _format_moscow(record.marked_at, with_seconds=True) if record.marked_at else "-"

                if status not in status_filter_set:
                    continue
                if source_key not in source_filter_set:
                    continue

                status_label = ATTENDANCE_STATUS_LABELS.get(status, status)
                export_rows.append(
                    {
                        "date": lesson_date_iso,
                        "day": day_label,
                        "pair": pair_label,
                        "time": pair_time or "-",
                        "course": course_title,
                        "group": group_name_map.get(int(group_id), f"Группа #{group_id}"),
                        "room": room_label,
                        "student": student.fio,
                        "status": status_label,
                        "presence": "Был" if status == ATTENDANCE_STATUS_PRESENT else "Не был",
                        "source": _source_label(source_key),
                        "ip": source_ip,
                        "marked_at": marked_at_display,
                    }
                )

        wb = Workbook()
        ws = wb.active
        ws.title = "Посещаемость"

        total_columns = 13
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        ws["A1"] = "Выгрузка посещаемости"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        selected_groups_label = (
            ", ".join(group_name_map[int(gid)] for gid in selected_group_ids)
            if selected_group_ids
            else "Все группы"
        )
        selected_courses_label = (
            ", ".join(course_title_map[int(cid)] for cid in selected_course_ids)
            if selected_course_ids
            else "Все предметы"
        )
        status_filter_label = ", ".join(ATTENDANCE_STATUS_LABELS.get(value, value) for value in selected_statuses)
        source_filter_label = ", ".join(_source_label(value) for value in selected_sources)
        if selected_student:
            student_filter_label = f"{selected_student.fio} ({group_name_map.get(selected_student_group_id, f'Группа #{selected_student_group_id}')})"
        elif student_query:
            student_filter_label = student_query
        else:
            student_filter_label = "Все студенты"

        ws["A2"] = f"Сформировано: {_format_moscow(datetime.now(timezone.utc), with_seconds=True)} (МСК)"
        ws["A3"] = f"Период: {date_from.isoformat()} - {date_to.isoformat()}"
        ws["A4"] = f"Группы: {selected_groups_label}"
        ws["A5"] = f"Предметы: {selected_courses_label}"
        ws["A6"] = f"Статусы: {status_filter_label} | Источники: {source_filter_label}"
        ws["A7"] = f"Студент: {student_filter_label}"

        for row_idx in range(2, 8):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_columns)
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        header_row = 8
        headers = [
            "Дата",
            "День",
            "Пара",
            "Время",
            "Предмет",
            "Группа",
            "Аудитория",
            "Студент",
            "Статус",
            "Факт",
            "Источник",
            "IP",
            "Отмечено (МСК)",
        ]
        header_fill = PatternFill("solid", fgColor="E9EEF5")
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        current_row = header_row + 1
        if export_rows:
            for item in export_rows:
                ws.cell(row=current_row, column=1, value=item["date"])
                ws.cell(row=current_row, column=2, value=item["day"])
                ws.cell(row=current_row, column=3, value=item["pair"])
                ws.cell(row=current_row, column=4, value=item["time"])
                ws.cell(row=current_row, column=5, value=item["course"])
                ws.cell(row=current_row, column=6, value=item["group"])
                ws.cell(row=current_row, column=7, value=item["room"])
                ws.cell(row=current_row, column=8, value=item["student"])
                ws.cell(row=current_row, column=9, value=item["status"])
                ws.cell(row=current_row, column=10, value=item["presence"])
                ws.cell(row=current_row, column=11, value=item["source"])
                ws.cell(row=current_row, column=12, value=item["ip"])
                ws.cell(row=current_row, column=13, value=item["marked_at"])
                current_row += 1
        else:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_columns)
            ws.cell(row=current_row, column=1, value="По выбранным фильтрам данные не найдены.")
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

        widths = {
            1: 12,
            2: 14,
            3: 11,
            4: 11,
            5: 34,
            6: 18,
            7: 13,
            8: 34,
            9: 21,
            10: 10,
            11: 16,
            12: 18,
            13: 22,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = f"A{header_row + 1}"
        if export_rows:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(total_columns)}{current_row - 1}"

        book_io = io.BytesIO()
        wb.save(book_io)
        book_io.seek(0)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename_raw = (
            f"journal_attendance_{date_from.isoformat()}_{date_to.isoformat()}_{timestamp}"
        )
        filename = f"{_safe_excel_filename(filename_raw)}.xlsx"
        return send_file(
            book_io,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
